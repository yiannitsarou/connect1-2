#!/usr/bin/env python3
"""
Unified Team Optimizer - Fill & Optimize in 1-Click
FIX v3.9 FINAL: Use .startswith() from working code + support Ν/N variants everywhere
"""
import streamlit as st
import openpyxl
from openpyxl.styles import Alignment, PatternFill, Font
from dataclasses import dataclass
from typing import Dict, List, Tuple, Optional
import io


@dataclass
class StudentData:
    """Δεδομένα μαθητή"""
    name: str
    gender: str
    teacher_child: str
    calm: str
    special_needs: str
    greek_knowledge: str
    friends: List[str]
    conflicts: int
    choice: int


@dataclass
class Student:
    """Student για optimizer"""
    name: str
    choice: int
    gender: str
    greek_knowledge: str
    friends: List[str]
    locked: bool


class UnifiedProcessor:
    """Ενοποιημένος processor: Fill + Optimize"""
    
    def __init__(self):
        self.students_data: Dict[str, StudentData] = {}
        self.teams_students: Dict[str, List[str]] = {}
        self.students: Dict[str, Student] = {}
        self.teams: Dict[str, List[str]] = {}
        self.target_ep3 = 3
        self.target_gender = 4
        self.target_greek = 4
    
    # ==================== PHASE 1: FILL EXCEL ====================
    
    def read_source_data(self, file_bytes: bytes) -> None:
        """Διάβασμα δεδομένων από Παράδειγμα1.xlsx"""
        wb = openpyxl.load_workbook(io.BytesIO(file_bytes), data_only=True)
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            headers = {}
            for col_idx, cell in enumerate(sheet[1], start=1):
                if cell.value:
                    header = str(cell.value).strip()
                    headers[header] = col_idx
            
            # DEBUG: Print all headers found
            st.write(f"🔍 **Sheet '{sheet_name}' Headers:**")
            st.write(headers)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            for row_idx in range(2, sheet.max_row + 1):
                name_cell = sheet.cell(row_idx, headers['ΟΝΟΜΑ'])
                name = name_cell.value
                
                if not name or str(name).strip() == '':
                    continue
                
                name = str(name).strip()
                
                def safe_get(header, default=''):
                    if header in headers:
                        col_idx = headers[header]
                        val = sheet.cell(row_idx, col_idx).value
                        if val is not None and str(val).strip() != '':
                            return str(val).strip()
                    return default
                
                friends_str = safe_get('ΦΙΛΟΙ', '')
                friends = [f.strip() for f in friends_str.split(',') if f.strip()] if friends_str else []
                
                choice_val = 1
                if 'ΕΠΙΔΟΣΗ' in headers:
                    epidosi_cell = sheet.cell(row_idx, headers['ΕΠΙΔΟΣΗ']).value
                    if epidosi_cell is not None:
                        try:
                            choice_val = int(epidosi_cell)
                        except:
                            choice_val = 1
                
                # FIX v3.9.1: Try multiple column name variants for Greek knowledge
                greek_raw = None
                found_column = None
                for possible_header in ['ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ ΓΝΩΣΗ ΕΛΛΗΝΙΚΩΝ', 
                                       'ΚΑΛΗ_ΓΝΩΣΗ', 'ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ']:
                    greek_raw = safe_get(possible_header, None)
                    if greek_raw is not None:
                        found_column = possible_header
                        break
                
                # DEBUG: Print which column was found (only first 3 students)
                if row_idx <= 4:
                    st.write(f"Row {row_idx} ({name}): Found column '{found_column}' with value '{greek_raw}'")
                
                if greek_raw is None or greek_raw == '':
                    greek_val = 'Ν'  # Default to ΝΑΙ only if empty
                else:
                    greek_str = str(greek_raw).strip().upper()
                    
                    # Use startswith() - πιάνει 'Ν', 'ΝΑΙ', 'N', etc.
                    if greek_str.startswith('Ν') or greek_str.startswith('N'):
                        greek_val = 'Ν'  # ΝΑΙ
                    elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                        greek_val = 'Ο'  # ΟΧΙ
                    else:
                        print(f"⚠️ Unknown ΚΑΛΗ_ΓΝΩΣΗ '{greek_raw}' for {name}, defaulting to Ν")
                        greek_val = 'Ν'
                
                self.students_data[name] = StudentData(
                    name=name,
                    gender=safe_get('ΦΥΛΟ', 'Κ'),
                    teacher_child=safe_get('ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ', 'Ο'),
                    calm=safe_get('ΖΩΗΡΟΣ', 'Ο'),
                    special_needs=safe_get('ΙΔΙΑΙΤΕΡΟΤΗΤΑ', 'Ο'),
                    greek_knowledge=greek_val,
                    friends=friends,
                    conflicts=0,
                    choice=choice_val
                )
        
        wb.close()
    
    def fill_target_excel(self, target_bytes: bytes) -> bytes:
        """Συμπλήρωση STEP7_FINAL_SCENARIO (in-memory)"""
        wb = openpyxl.load_workbook(io.BytesIO(target_bytes))
        
        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
            self._fill_sheet(sheet, sheet_name)
        
        self._create_categorization_sheet(wb)
        
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        
        return output.getvalue()
    
    def _fill_sheet(self, sheet, team_name: str) -> int:
        """Συμπλήρωση ενός sheet"""
        headers_map = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                header = str(cell.value).strip().upper()
                header_key = header.replace('_', '').replace(' ', '')
                headers_map[header_key] = col_idx
        
        if 'ΟΝΟΜΑ' not in headers_map:
            return 0
        
        required_headers = ['ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        next_col = max(headers_map.values()) + 1
        
        for req_header in required_headers:
            if req_header not in headers_map:
                cell = sheet.cell(1, next_col)
                original_header = req_header.replace('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ')
                cell.value = original_header
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.font = Font(bold=True)
                headers_map[req_header] = next_col
                next_col += 1
        
        filled_count = 0
        
        if team_name not in self.teams_students:
            self.teams_students[team_name] = []
        
        for row_idx in range(2, sheet.max_row + 1):
            name_cell = sheet.cell(row_idx, headers_map['ΟΝΟΜΑ'])
            name = name_cell.value
            
            if not name or str(name).strip() == '':
                continue
            
            name = str(name).strip()
            
            if name not in self.students_data:
                continue
            
            student_data = self.students_data[name]
            self.teams_students[team_name].append(name)
            
            if 'ΦΥΛΟ' in headers_map:
                col = headers_map['ΦΥΛΟ']
                sheet.cell(row_idx, col).value = student_data.gender
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            if 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ' in headers_map:
                col = headers_map['ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ']
                sheet.cell(row_idx, col).value = student_data.greek_knowledge
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            for key in headers_map.keys():
                if 'ΚΑΛΗ' in key and 'ΓΝΩΣΗ' in key and 'ΕΛΛΗΝΙΚΩΝ' in key:
                    col = headers_map[key]
                    sheet.cell(row_idx, col).value = student_data.greek_knowledge
                    sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
                    break
            
            if 'ΦΙΛΟΙ' in headers_map:
                col = headers_map['ΦΙΛΟΙ']
                sheet.cell(row_idx, col).value = ', '.join(student_data.friends) if student_data.friends else ''
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='left', vertical='center')
            
            if 'ΕΠΙΔΟΣΗ' in headers_map:
                col = headers_map['ΕΠΙΔΟΣΗ']
                sheet.cell(row_idx, col).value = student_data.choice
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            filled_count += 1
        
        return filled_count
    
    def _create_categorization_sheet(self, workbook) -> None:
        """Δημιουργία sheet ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ"""
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in workbook.sheetnames:
            del workbook['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ']
        
        cat_sheet = workbook.create_sheet('ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ')
        
        headers = ['ΜΑΘΗΤΗΣ Α', 'ΜΑΘΗΤΗΣ Β', 'ΚΑΤΗΓΟΡΙΑ ΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ', 'LOCKED', 'ΤΜΗΜΑ']
        for col_idx, header in enumerate(headers, start=1):
            cell = cat_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        all_students = []
        for team_name in sorted(self.teams_students.keys()):
            for student_name in self.teams_students[team_name]:
                if student_name in self.students_data:
                    student = self.students_data[student_name]
                    all_students.append({
                        'name': student_name,
                        'data': student,
                        'team': team_name
                    })
        
        row_idx = 2
        processed = set()
        
        for i, student_a in enumerate(all_students):
            if student_a['name'] in processed:
                continue
            
            for j, student_b in enumerate(all_students[i+1:], start=i+1):
                if student_b['name'] in processed:
                    continue
                
                if (student_b['name'] in student_a['data'].friends or 
                    student_a['name'] in student_b['data'].friends):
                    
                    category = self._determine_category(
                        student_a['data'].gender,
                        student_a['data'].greek_knowledge,
                        student_b['data'].gender,
                        student_b['data'].greek_knowledge
                    )
                    
                    epidosi_text = f"{student_a['data'].choice}, {student_b['data'].choice}"
                    
                    cat_sheet.cell(row_idx, 1).value = student_a['name']
                    cat_sheet.cell(row_idx, 2).value = student_b['name']
                    cat_sheet.cell(row_idx, 3).value = category
                    cat_sheet.cell(row_idx, 4).value = epidosi_text
                    
                    is_locked = (self._is_student_locked(student_a['data']) or 
                                 self._is_student_locked(student_b['data']))
                    cat_sheet.cell(row_idx, 5).value = 'LOCKED' if is_locked else 'ΟΧΙ'
                    
                    if is_locked:
                        team_text = 'LOCKED'
                    else:
                        team_text = f"{student_a['team']},{student_b['team']}"
                    cat_sheet.cell(row_idx, 6).value = team_text
                    
                    for col in range(1, 7):
                        cat_sheet.cell(row_idx, col).alignment = Alignment(
                            horizontal='left' if col <= 2 else 'center',
                            vertical='center'
                        )
                    
                    processed.add(student_a['name'])
                    processed.add(student_b['name'])
                    row_idx += 1
                    break
        
        cat_sheet.column_dimensions['A'].width = 30
        cat_sheet.column_dimensions['B'].width = 30
        cat_sheet.column_dimensions['C'].width = 35
        cat_sheet.column_dimensions['D'].width = 12
        cat_sheet.column_dimensions['E'].width = 12
        cat_sheet.column_dimensions['F'].width = 20
        
        self._create_single_sheet(workbook, all_students, processed)
    
    def _is_student_locked(self, student: StudentData) -> bool:
        """
        FIX v3.5 CRITICAL: Ελέγχουμε ΖΩΗΡΟΣ, ΠΑΙΔΙ_ΕΚΠΑΙΔΕΥΤΙΚΟΥ, ΙΔΙΑΙΤΕΡΟΤΗΤΑ
        Ν = ΝΑΙ (locked) σε αυτά τα πεδία
        Η ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ ΔΕΝ είναι locked field!
        """
        return (student.calm == 'Ν' or 
                student.teacher_child == 'Ν' or 
                student.special_needs == 'Ν')
    
    def _determine_category(self, gender_a: str, greek_a: str, gender_b: str, greek_b: str) -> str:
        """Καθορισμός κατηγορίας δυάδας"""
        if gender_a != gender_b:
            return "Ομάδες Μικτού Φύλου"
        
        gender_label = "Κορίτσια" if gender_a == "Κ" else "Αγόρια"
        
        if greek_a == greek_b:
            if greek_a == "Ν":
                return f"Καλή Γνώση ({gender_label})"
            else:
                return f"όχι Καλή Γνώση ({gender_label})"
        else:
            return f"Μικτής Γνώσης ({gender_label})"
    
    def _determine_single_category(self, gender: str, greek_knowledge: str) -> str:
        """Καθορισμός κατηγορίας για μεμονωμένο μαθητή"""
        gender_label = "Κορίτσια" if gender == "Κ" else "Αγόρια"
        
        if greek_knowledge == "Ν":
            return f"{gender_label} - Ν (Καλή γνώση)"
        else:
            return f"{gender_label} - Ο (όχι καλή γνώση)"
    
    def _create_single_sheet(self, workbook, all_students: List[Dict], processed_names: set) -> None:
        """Δημιουργία sheet SINGLE"""
        if 'SINGLE' in workbook.sheetnames:
            del workbook['SINGLE']
        
        single_sheet = workbook.create_sheet('SINGLE')
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΚΑΤΗΓΟΡΙΑ SINGLE', 'ΤΜΗΜΑ', 'LOCKED']
        for col_idx, header in enumerate(headers, start=1):
            cell = single_sheet.cell(1, col_idx)
            cell.value = header
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        single_students = []
        for student in all_students:
            if student['name'] not in processed_names:
                single_students.append(student)
        
        single_students.sort(key=lambda x: x['name'])
        
        for row_idx, student in enumerate(single_students, start=2):
            student_data = student['data']
            team_name = student['team']
            
            single_sheet.cell(row_idx, 1).value = student['name']
            single_sheet.cell(row_idx, 2).value = student_data.gender
            single_sheet.cell(row_idx, 3).value = student_data.greek_knowledge
            single_sheet.cell(row_idx, 4).value = student_data.choice
            
            category = self._determine_single_category(student_data.gender, student_data.greek_knowledge)
            single_sheet.cell(row_idx, 5).value = category
            
            is_locked = self._is_student_locked(student_data)
            single_sheet.cell(row_idx, 7).value = 'LOCKED' if is_locked else 'ΟΧΙ'
            
            if is_locked:
                single_sheet.cell(row_idx, 6).value = 'LOCKED'
            else:
                single_sheet.cell(row_idx, 6).value = team_name
            
            for col in range(1, 8):
                single_sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col == 1 else 'center',
                    vertical='center'
                )
        
        single_sheet.column_dimensions['A'].width = 30
        single_sheet.column_dimensions['B'].width = 12
        single_sheet.column_dimensions['C'].width = 25
        single_sheet.column_dimensions['D'].width = 12
        single_sheet.column_dimensions['E'].width = 35
        single_sheet.column_dimensions['F'].width = 20
        single_sheet.column_dimensions['G'].width = 12
    
    # ==================== PHASE 2: OPTIMIZE ====================
    
    def load_filled_data(self, filled_bytes: bytes) -> None:
        """Φόρτωση δεδομένων από filled Excel για optimization"""
        wb = openpyxl.load_workbook(io.BytesIO(filled_bytes), data_only=True)
        
        if 'ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ' in wb.sheetnames:
            self._load_from_kategoriopoihsh(wb['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ'])
        
        if 'SINGLE' in wb.sheetnames:
            self._load_from_single(wb['SINGLE'])
        
        for sheet_name in wb.sheetnames:
            if sheet_name in ['ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ', 'SINGLE']:
                continue
            
            sheet = wb[sheet_name]
            headers = self._parse_headers(sheet)
            
            if 'ΟΝΟΜΑ' not in headers:
                continue
            
            self.teams[sheet_name] = []
            
            for row_idx in range(2, sheet.max_row + 1):
                name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
                if name and name in self.students:
                    self.teams[sheet_name].append(name)
        
        wb.close()
    
    def _load_from_kategoriopoihsh(self, sheet) -> None:
        """Διάβασμα δυάδων από ΚΑΤΗΓΟΡΙΟΠΟΙΗΣΗ sheet"""
        headers = self._parse_headers(sheet)
        
        required = ['ΜΑΘΗΤΗΣΑ', 'ΜΑΘΗΤΗΣΒ', 'ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            name_a = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΑ'))
            name_b = self._get_cell_value(sheet, row_idx, headers.get('ΜΑΘΗΤΗΣΒ'))
            category = self._get_cell_value(sheet, row_idx, headers.get('ΚΑΤΗΓΟΡΙΑΔΥΑΔΑΣ'))
            epidosh_raw = self._get_cell_value(sheet, row_idx, headers.get('ΕΠΙΔΟΣΗ'))
            locked_val = self._get_cell_value(sheet, row_idx, headers.get('LOCKED'))
            
            if not name_a or not name_b or not category:
                continue
            
            epidosh_a, epidosh_b = 1, 1
            if ',' in epidosh_raw:
                parts = epidosh_raw.split(',')
                try:
                    epidosh_a = int(parts[0].strip())
                    epidosh_b = int(parts[1].strip())
                except:
                    pass
            
            gender_a = gender_b = 'Α'
            greek_a = greek_b = 'Ν'
            
            if 'Αγόρια' in category or 'Αγόρ' in category:
                gender_a = gender_b = 'Α'
            elif 'Κορίτσια' in category or 'Κορίτ' in category:
                gender_a = gender_b = 'Κ'
            
            if 'όχι Καλή Γνώση' in category or 'όχι καλή' in category.lower():
                greek_a = greek_b = 'Ο'
            elif 'Καλή Γνώση' in category or 'Καλή γνώση' in category:
                greek_a = greek_b = 'Ν'
            elif 'Μικτής' in category or 'μικτής' in category.lower():
                greek_a = greek_b = 'Ν'
            
            is_locked = (locked_val == 'LOCKED')
            
            if name_a not in self.students:
                self.students[name_a] = Student(
                    name=name_a,
                    choice=epidosh_a,
                    gender=gender_a,
                    greek_knowledge=greek_a,
                    friends=[name_b],
                    locked=is_locked
                )
            
            if name_b not in self.students:
                self.students[name_b] = Student(
                    name=name_b,
                    choice=epidosh_b,
                    gender=gender_b,
                    greek_knowledge=greek_b,
                    friends=[name_a],
                    locked=is_locked
                )
    
    def _load_from_single(self, sheet) -> None:
        """Διάβασμα μονών μαθητών από SINGLE sheet"""
        headers = self._parse_headers(sheet)
        
        required = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ']
        missing = [h for h in required if h not in headers]
        if missing:
            return
        
        for row_idx in range(2, sheet.max_row + 1):
            name = self._get_cell_value(sheet, row_idx, headers.get('ΟΝΟΜΑ'))
            if not name:
                continue
            
            if name in self.students:
                continue
            
            gender_col = headers.get('ΦΥΛΟ')
            greek_col = (headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ') or 
                        headers.get('ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ') or
                        headers.get('ΚΑΛΗΓΝΩΣΗΕΛΛΗΝΙΚΩΝ'))
            epidosh_col = headers.get('ΕΠΙΔΟΣΗ')
            locked_col = headers.get('LOCKED')
            
            gender = self._get_cell_value(sheet, row_idx, gender_col, 'Α')
            
            # Greek knowledge - use startswith() like working code
            raw_greek = sheet.cell(row_idx, greek_col).value if greek_col else 'Ν'
            if raw_greek:
                greek_str = str(raw_greek).strip().upper()
                if greek_str.startswith('Ν') or greek_str.startswith('N'):
                    greek = 'Ν'
                elif greek_str.startswith('Ο') or greek_str.startswith('O'):
                    greek = 'Ο'
                else:
                    greek = 'Ν'
            else:
                greek = 'Ν'
            
            raw_epidosh = sheet.cell(row_idx, epidosh_col).value if epidosh_col else 1
            try:
                epidosh = int(raw_epidosh) if raw_epidosh else 1
            except:
                epidosh = 1
            
            locked_val = self._get_cell_value(sheet, row_idx, locked_col)
            is_locked = (locked_val == 'LOCKED')
            
            self.students[name] = Student(
                name=name,
                choice=epidosh,
                gender=gender,
                greek_knowledge=greek,
                friends=[],
                locked=is_locked
            )
    
    def _parse_headers(self, sheet) -> Dict[str, int]:
        """Normalization headers"""
        headers = {}
        for col_idx, cell in enumerate(sheet[1], start=1):
            if cell.value:
                raw_header = str(cell.value).strip()
                headers[raw_header] = col_idx
                normalized = raw_header.upper().replace(' ', '').replace('_', '')
                headers[normalized] = col_idx
        return headers
    
    def _get_cell_value(self, sheet, row: int, col: int, default=''):
        if col is None:
            return default
        val = sheet.cell(row, col).value
        return str(val).strip() if val is not None else default
    
    def calculate_spreads(self) -> Dict[str, int]:
        """Υπολογισμός spreads"""
        stats = self._get_team_stats()
        if not stats:
            return {'ep3': 0, 'boys': 0, 'girls': 0, 'greek_yes': 0}
        
        ep3_vals = [s['ep3'] for s in stats.values()]
        boys_vals = [s['boys'] for s in stats.values()]
        girls_vals = [s['girls'] for s in stats.values()]
        greek_yes_vals = [s['greek_yes'] for s in stats.values()]
        
        return {
            'ep3': max(ep3_vals) - min(ep3_vals),
            'boys': max(boys_vals) - min(boys_vals),
            'girls': max(girls_vals) - min(girls_vals),
            'greek_yes': max(greek_yes_vals) - min(greek_yes_vals)
        }
    
    def _get_team_stats(self) -> Dict:
        """Μέτρηση stats ανά τμήμα"""
        stats = {}
        for team_name, student_names in self.teams.items():
            boys = girls = greek_yes = greek_no = ep1 = ep2 = ep3 = 0
            
            for name in student_names:
                if name not in self.students:
                    continue
                s = self.students[name]
                
                if s.gender == 'Α':
                    boys += 1
                elif s.gender == 'Κ':
                    girls += 1
                
                # FIX v3.9: Support BOTH Greek Ν (U+039D) and Latin N (U+004E)
                if s.greek_knowledge in ['Ν', 'N']:
                    greek_yes += 1
                elif s.greek_knowledge in ['Ο', 'O']:
                    greek_no += 1
                
                if s.choice == 1:
                    ep1 += 1
                elif s.choice == 2:
                    ep2 += 1
                elif s.choice == 3:
                    ep3 += 1
            
            stats[team_name] = {
                'boys': boys, 'girls': girls,
                'greek_yes': greek_yes, 'greek_no': greek_no,
                'ep1': ep1, 'ep2': ep2, 'ep3': ep3
            }
        
        return stats
    
    def optimize(self, max_iterations: int = 100) -> Tuple[List[Dict], Dict]:
        """Asymmetric optimization"""
        applied_swaps = []
        
        for iteration in range(max_iterations):
            spreads = self.calculate_spreads()
            
            if (spreads['ep3'] <= self.target_ep3 and
                spreads['boys'] <= self.target_gender and
                spreads['girls'] <= self.target_gender and
                spreads['greek_yes'] <= self.target_greek):
                break
            
            stats = self._get_team_stats()
            ep3_counts = {team: stats[team]['ep3'] for team in stats.keys()}
            
            max_team = max(ep3_counts.items(), key=lambda x: x[1])[0]
            min_team = min(ep3_counts.items(), key=lambda x: x[1])[0]
            
            if ep3_counts[max_team] - ep3_counts[min_team] <= self.target_ep3:
                break
            
            all_swaps = self._generate_asymmetric_swaps(max_team, min_team)
            
            if not all_swaps:
                break
            
            best_swap = self._select_best_swap(all_swaps)
            
            if not best_swap:
                break
            
            self._apply_swap(best_swap)
            applied_swaps.append(best_swap)
        
        final_spreads = self.calculate_spreads()
        return applied_swaps, final_spreads
    
    def _generate_asymmetric_swaps(self, max_team: str, min_team: str) -> List[Dict]:
        """Γέννηση asymmetric swaps"""
        swaps = []
        
        max_solos_ep3 = self._get_solos_with_ep3(max_team)
        max_pairs_ep3 = self._get_pairs_with_ep3(max_team)
        min_solos_non_ep3 = self._get_solos_without_ep3(min_team)
        min_pairs_non_ep3 = self._get_pairs_without_ep3(min_team)
        
        # P1: Solo(ep3)↔Solo(ep1/2) - same gender + greek
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if (solo_max['student'].gender == solo_min['student'].gender and
                    solo_max['student'].greek_knowledge == solo_min['student'].greek_knowledge):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P1',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 1
                        })
        
        # P2: Pair swaps
        for pair_max in max_pairs_ep3:
            for pair_min in min_pairs_non_ep3:
                if (pair_max['student_a'].gender == pair_min['student_a'].gender and
                    pair_max['student_b'].gender == pair_min['student_b'].gender and
                    pair_max['student_a'].greek_knowledge == pair_min['student_a'].greek_knowledge and
                    pair_max['student_b'].greek_knowledge == pair_min['student_b'].greek_knowledge):
                    
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [pair_max['name_a'], pair_max['name_b']],
                        min_team, [pair_min['name_a'], pair_min['name_b']]
                    )
                    
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Pair(ep3+X)↔Pair(ep1/2)-P2',
                            'from_team': max_team,
                            'students_out': [pair_max['name_a'], pair_max['name_b']],
                            'to_team': min_team,
                            'students_in': [pair_min['name_a'], pair_min['name_b']],
                            'improvement': improvement,
                            'priority': 2
                        })
        
        # P3: Relaxed (only gender match)
        for solo_max in max_solos_ep3:
            for solo_min in min_solos_non_ep3:
                if solo_max['student'].gender == solo_min['student'].gender:
                    improvement = self._calc_asymmetric_improvement(
                        max_team, [solo_max['name']],
                        min_team, [solo_min['name']]
                    )
                    if improvement['improves']:
                        swaps.append({
                            'type': 'Solo(ep3)↔Solo(ep1/2)-P3',
                            'from_team': max_team,
                            'students_out': [solo_max['name']],
                            'to_team': min_team,
                            'students_in': [solo_min['name']],
                            'improvement': improvement,
                            'priority': 3
                        })
        
        return swaps
    
    def _get_solos_with_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice != 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_with_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    if student_a.choice == 3 or student_b.choice == 3:
                        pairs.append({
                            'name_a': name_a, 'name_b': name_b,
                            'student_a': student_a, 'student_b': student_b,
                            'ep_combo': f"{student_a.choice},{student_b.choice}"
                        })
                        processed.add(name_a)
                        processed.add(name_b)
                        break
        return pairs
    
    def _get_solos_without_ep3(self, team_name: str) -> List[Dict]:
        solos = []
        student_names = self.teams[team_name]
        for name in student_names:
            if name not in self.students:
                continue
            student = self.students[name]
            if student.locked or student.choice == 3:
                continue
            has_friend = any(f in student_names for f in student.friends)
            if not has_friend:
                solos.append({'name': name, 'student': student})
        return solos
    
    def _get_pairs_without_ep3(self, team_name: str) -> List[Dict]:
        pairs = []
        processed = set()
        student_names = self.teams[team_name]
        for name_a in student_names:
            if name_a in processed or name_a not in self.students:
                continue
            student_a = self.students[name_a]
            if student_a.locked:
                continue
            for name_b in student_names:
                if name_b == name_a or name_b in processed or name_b not in self.students:
                    continue
                student_b = self.students[name_b]
                if student_b.locked:
                    continue
                if name_b in student_a.friends or name_a in student_b.friends:
                    if student_a.choice != 3 and student_b.choice != 3:
                        pairs.append({
                            'name_a': name_a, 'name_b': name_b,
                            'student_a': student_a, 'student_b': student_b,
                            'ep_combo': f"{student_a.choice},{student_b.choice}"
                        })
                        processed.add(name_a)
                        processed.add(name_b)
                        break
        return pairs
    
    def _calc_asymmetric_improvement(self, team_high: str, names_out: List[str],
                                      team_low: str, names_in: List[str]) -> Dict:
        """FIX v3.9: Support Ν/N variants in improvement calculation"""
        stats_before = self._get_team_stats()
        stats_after = {k: v.copy() for k, v in stats_before.items()}
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_high]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] -= 1
                # FIX: Support both Ν and N
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_high]['greek_yes'] -= 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_high]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_high]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_high]['girls'] += 1
                # FIX: Support both Ν and N
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_high]['greek_yes'] += 1
        
        for name in names_in:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] -= 1
                if s.gender == 'Α': stats_after[team_low]['boys'] -= 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] -= 1
                # FIX: Support both Ν and N
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_low]['greek_yes'] -= 1
        
        for name in names_out:
            if name in self.students:
                s = self.students[name]
                if s.choice == 3: stats_after[team_low]['ep3'] += 1
                if s.gender == 'Α': stats_after[team_low]['boys'] += 1
                elif s.gender == 'Κ': stats_after[team_low]['girls'] += 1
                # FIX: Support both Ν and N
                if s.greek_knowledge in ['Ν', 'N']: stats_after[team_low]['greek_yes'] += 1
        
        ep3_before = max(s['ep3'] for s in stats_before.values()) - min(s['ep3'] for s in stats_before.values())
        ep3_after = max(s['ep3'] for s in stats_after.values()) - min(s['ep3'] for s in stats_after.values())
        
        boys_before = max(s['boys'] for s in stats_before.values()) - min(s['boys'] for s in stats_before.values())
        boys_after = max(s['boys'] for s in stats_after.values()) - min(s['boys'] for s in stats_after.values())
        
        girls_before = max(s['girls'] for s in stats_before.values()) - min(s['girls'] for s in stats_before.values())
        girls_after = max(s['girls'] for s in stats_after.values()) - min(s['girls'] for s in stats_after.values())
        
        greek_before = max(s['greek_yes'] for s in stats_before.values()) - min(s['greek_yes'] for s in stats_before.values())
        greek_after = max(s['greek_yes'] for s in stats_after.values()) - min(s['greek_yes'] for s in stats_after.values())
        
        delta_ep3 = ep3_before - ep3_after
        delta_boys = boys_before - boys_after
        delta_girls = girls_before - girls_after
        delta_greek = greek_before - greek_after
        
        improves = delta_ep3 > 0 or (delta_ep3 == 0 and (delta_boys > 0 or delta_girls > 0 or delta_greek > 0))
        
        return {
            'improves': improves,
            'delta_ep3': delta_ep3,
            'delta_boys': delta_boys,
            'delta_girls': delta_girls,
            'delta_greek': delta_greek,
            'ep3_before': ep3_before,
            'ep3_after': ep3_after
        }
    
    def _select_best_swap(self, swaps: List[Dict]) -> Optional[Dict]:
        if not swaps:
            return None
        
        swaps.sort(
            key=lambda x: (
                -x['improvement']['delta_ep3'],
                -(x['improvement']['delta_boys'] + x['improvement']['delta_girls']),
                -x['improvement']['delta_greek'],
                x['priority']
            )
        )
        
        return swaps[0]
    
    def _apply_swap(self, swap: Dict) -> None:
        from_team = swap['from_team']
        to_team = swap['to_team']
        students_out = swap['students_out']
        students_in = swap['students_in']
        
        for name in students_out:
            if name in self.teams[from_team]:
                self.teams[from_team].remove(name)
        
        for name in students_in:
            if name in self.teams[to_team]:
                self.teams[to_team].remove(name)
        
        for name in students_out:
            self.teams[to_team].append(name)
        
        for name in students_in:
            self.teams[from_team].append(name)
    
    def export_optimized_excel(self, applied_swaps: List[Dict], final_spreads: Dict) -> bytes:
        """Εξαγωγή optimized Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        for team_name in sorted(self.teams.keys()):
            self._create_team_sheet(wb, team_name)
        
        self._create_statistics_sheet(wb, final_spreads)
        self._create_swaps_log_sheet(wb, applied_swaps)
        
        output = io.BytesIO()
        wb.save(output)
        wb.close()
        output.seek(0)
        
        return output.getvalue()
    
    def _create_team_sheet(self, wb, team_name: str) -> None:
        sheet = wb.create_sheet(team_name)
        
        headers = ['ΟΝΟΜΑ', 'ΦΥΛΟ', 'ΚΑΛΗ_ΓΝΩΣΗ_ΕΛΛΗΝΙΚΩΝ', 'ΕΠΙΔΟΣΗ', 'ΦΙΛΟΙ']
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='DDEBF7', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        row_idx = 2
        for name in sorted(self.teams[team_name]):
            if name not in self.students:
                continue
            
            student = self.students[name]
            
            # FIX v3.7: Normalize greek_knowledge to Greek chars
            greek_val = student.greek_knowledge
            if greek_val in ['N', 'n']:
                greek_val = 'Ν'  # Force Greek Nu
            elif greek_val in ['O', 'o']:
                greek_val = 'Ο'  # Force Greek Omicron
            
            sheet.cell(row_idx, 1).value = student.name
            sheet.cell(row_idx, 2).value = student.gender
            sheet.cell(row_idx, 3).value = greek_val  # Use normalized value
            sheet.cell(row_idx, 4).value = student.choice
            sheet.cell(row_idx, 5).value = ', '.join(student.friends)
            
            for col in range(1, 6):
                sheet.cell(row_idx, col).alignment = Alignment(
                    horizontal='left' if col in [1,5] else 'center', 
                    vertical='center'
                )
            
            row_idx += 1
        
        sheet.column_dimensions['A'].width = 30
        sheet.column_dimensions['B'].width = 12
        sheet.column_dimensions['C'].width = 25
        sheet.column_dimensions['D'].width = 12
        sheet.column_dimensions['E'].width = 40
    
    def _create_statistics_sheet(self, wb, spreads: Dict) -> None:
        sheet = wb.create_sheet('ΒΕΛΤΙΩΜΕΝΗ_ΣΤΑΤΙΣΤΙΚΗ')
        
        headers = ['Τμήμα', 'Σύνολο', 'Αγόρια', 'Κορίτσια', 
                   'Γνώση (ΝΑΙ)', 'Γνώση (ΟΧΙ)', 'Επ1', 'Επ2', 'Επ3']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='C6E0B4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        stats = self._get_team_stats()
        row_idx = 2
        for team_name in sorted(self.teams.keys()):
            if team_name not in stats:
                continue
            s = stats[team_name]
            
            sheet.cell(row_idx, 1).value = team_name
            sheet.cell(row_idx, 2).value = len(self.teams[team_name])
            sheet.cell(row_idx, 3).value = s['boys']
            sheet.cell(row_idx, 4).value = s['girls']
            sheet.cell(row_idx, 5).value = s['greek_yes']
            sheet.cell(row_idx, 6).value = s['greek_no']
            sheet.cell(row_idx, 7).value = s['ep1']
            sheet.cell(row_idx, 8).value = s['ep2']
            sheet.cell(row_idx, 9).value = s['ep3']
            
            for col in range(1, 10):
                sheet.cell(row_idx, col).alignment = Alignment(horizontal='center', vertical='center')
            
            row_idx += 1
        
        row_idx += 2
        sheet.cell(row_idx, 1).value = 'ΤΕΛΙΚΑ SPREADS'
        sheet.cell(row_idx, 1).font = Font(bold=True, size=12)
        row_idx += 1
        
        summary_headers = ['Μετρική', 'Spread', 'Στόχος', 'Status']
        for col_idx, header in enumerate(summary_headers, start=1):
            cell = sheet.cell(row_idx, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='FFF2CC', fill_type='solid')
        row_idx += 1
        
        summary_data = [
            ('Spread Επίδοσης 3', spreads['ep3'], '≤ 3', '✅' if spreads['ep3'] <= 3 else '❌'),
            ('Spread Αγοριών', spreads['boys'], '≤ 4', '✅' if spreads['boys'] <= 4 else '❌'),
            ('Spread Κοριτσιών', spreads['girls'], '≤ 4', '✅' if spreads['girls'] <= 4 else '❌'),
            ('Spread Γνώσης', spreads['greek_yes'], '≤ 4', '✅' if spreads['greek_yes'] <= 4 else '❌')
        ]
        
        for label, value, target, status in summary_data:
            sheet.cell(row_idx, 1).value = label
            sheet.cell(row_idx, 2).value = value
            sheet.cell(row_idx, 3).value = target
            sheet.cell(row_idx, 4).value = status
            
            if '✅' in status:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='C6EFCE', fill_type='solid')
            else:
                sheet.cell(row_idx, 2).fill = PatternFill(start_color='FFC7CE', fill_type='solid')
            
            row_idx += 1
        
        for col in ['A', 'B', 'C', 'D']:
            sheet.column_dimensions[col].width = 20
    
    def _create_swaps_log_sheet(self, wb, swaps: List[Dict]) -> None:
        sheet = wb.create_sheet('ΕΦΑΡΜΟΣΜΕΝΑ_SWAPS')
        
        headers = ['#', 'Τύπος', 'Από Τμήμα', 'Μαθητές OUT', 
                   'Προς Τμήμα', 'Μαθητές IN', 'Δ_ep3', 'Δ_φύλου', 'Δ_γνώσης', 'Priority']
        
        for col_idx, header in enumerate(headers, start=1):
            cell = sheet.cell(1, col_idx)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='D9E1F2', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        
        for idx, swap in enumerate(swaps, start=1):
            imp = swap['improvement']
            
            sheet.cell(idx + 1, 1).value = idx
            sheet.cell(idx + 1, 2).value = swap['type']
            sheet.cell(idx + 1, 3).value = swap['from_team']
            sheet.cell(idx + 1, 4).value = ', '.join(swap['students_out'])
            sheet.cell(idx + 1, 5).value = swap['to_team']
            sheet.cell(idx + 1, 6).value = ', '.join(swap['students_in'])
            sheet.cell(idx + 1, 7).value = f"+{imp['delta_ep3']}" if imp['delta_ep3'] > 0 else str(imp['delta_ep3'])
            sheet.cell(idx + 1, 8).value = f"+{imp['delta_boys'] + imp['delta_girls']}" if imp['delta_boys'] + imp['delta_girls'] > 0 else str(imp['delta_boys'] + imp['delta_girls'])
            sheet.cell(idx + 1, 9).value = f"+{imp['delta_greek']}" if imp['delta_greek'] > 0 else str(imp['delta_greek'])
            sheet.cell(idx + 1, 10).value = swap['priority']
            
            for col in range(1, 11):
                sheet.cell(idx + 1, col).alignment = Alignment(horizontal='center', vertical='center')
        
        for col, width in [('A',8),('B',25),('C',15),('D',35),('E',15),('F',35),('G',10),('H',10),('I',10),('J',10)]:
            sheet.column_dimensions[col].width = width


def main():
    st.set_page_config(
        page_title="Unified Team Optimizer",
        page_icon="🎯",
        layout="wide"
    )
    
    st.title("🎯 Unified Team Optimizer v3.9 FINAL")
    st.markdown("---")
    
    with st.expander("📖 Οδηγίες Χρήσης", expanded=False):
        st.markdown("""
        **FIX v3.9 FINAL:** 
        - ✅ Use .startswith() όπως working code (robust για ΝΑΙ/ΟΧΙ)
        - ✅ Support Ν/N variants σε ΟΛΑ τα σημεία (_get_team_stats + _calc_asymmetric_improvement)
        
        **Workflow:**
        1. Ανέβασε **Παράδειγμα1.xlsx** (δεδομένα μαθητών)
        2. Ανέβασε **STEP7_TEMPLATE.xlsx** (template τμημάτων)
        3. Πάτα "⚡ Fill & Optimize"
        4. Κατέβασε **ΒΕΛΤΙΩΜΕΝΗ_ΚΑΤΑΝΟΜΗ.xlsx**
        
        **Στόχοι:** Spread Επ3 ≤3, Φύλου ≤4, Γνώσης ≤4
        """)
    
    col1, col2 = st.columns(2)
    
    with col1:
        st.subheader("📥 Πηγή Δεδομένων")
        source_file = st.file_uploader("Ανέβασε Παράδειγμα1.xlsx", type=['xlsx'], key='source')
        if source_file:
            st.success(f"✅ {source_file.name}")
    
    with col2:
        st.subheader("📄 Template")
        template_file = st.file_uploader("Ανέβασε STEP7_TEMPLATE.xlsx", type=['xlsx'], key='template')
        if template_file:
            st.success(f"✅ {template_file.name}")
    
    st.markdown("---")
    
    if source_file and template_file:
        if st.button("⚡ Fill & Optimize", type="primary", use_container_width=True):
            with st.spinner("🔄 Phase 1/2: Filling..."):
                try:
                    processor = UnifiedProcessor()
                    source_bytes = source_file.read()
                    template_bytes = template_file.read()
                    
                    processor.read_source_data(source_bytes)
                    st.success(f"✅ Βρέθηκαν {len(processor.students_data)} μαθητές")
                    
                    filled_bytes = processor.fill_target_excel(template_bytes)
                    st.success("✅ Excel συμπληρώθηκε")
                    
                except Exception as e:
                    st.error(f"❌ Σφάλμα Phase 1: {str(e)}")
                    st.stop()
            
            with st.spinner("🔄 Phase 2/2: Optimizing..."):
                try:
                    processor.load_filled_data(filled_bytes)
                    spreads_before = processor.calculate_spreads()
                    
                    st.info("📊 **ΠΡΙΝ:**")
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Spread Επ3", spreads_before['ep3'])
                    with col2:
                        st.metric("Spread Αγόρια", spreads_before['boys'])
                    with col3:
                        st.metric("Spread Κορίτσια", spreads_before['girls'])
                    with col4:
                        st.metric("Spread Γνώση", spreads_before['greek_yes'])
                    
                    applied_swaps, spreads_after = processor.optimize(max_iterations=100)
                    
                    st.markdown("---")
                    st.success("✅ **ΜΕΤΑ:**")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Spread Επ3", spreads_after['ep3'], 
                                 delta=-(spreads_before['ep3'] - spreads_after['ep3']), delta_color="inverse")
                        if spreads_after['ep3'] <= 3:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤3")
                    
                    with col2:
                        st.metric("Spread Αγόρια", spreads_after['boys'],
                                 delta=-(spreads_before['boys'] - spreads_after['boys']), delta_color="inverse")
                        if spreads_after['boys'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤4")
                    
                    with col3:
                        st.metric("Spread Κορίτσια", spreads_after['girls'],
                                 delta=-(spreads_before['girls'] - spreads_after['girls']), delta_color="inverse")
                        if spreads_after['girls'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤4")
                    
                    with col4:
                        st.metric("Spread Γνώση", spreads_after['greek_yes'],
                                 delta=-(spreads_before['greek_yes'] - spreads_after['greek_yes']), delta_color="inverse")
                        if spreads_after['greek_yes'] <= 4:
                            st.success("✅")
                        else:
                            st.warning("⚠️ ≤4")
                    
                    st.markdown("---")
                    st.info(f"🔄 **{len(applied_swaps)} swaps εφαρμόστηκαν**")
                    
                    output_bytes = processor.export_optimized_excel(applied_swaps, spreads_after)
                    
                    st.download_button(
                        label="📥 Κατέβασε ΒΕΛΤΙΩΜΕΝΗ_ΚΑΤΑΝΟΜΗ.xlsx",
                        data=output_bytes,
                        file_name="ΒΕΛΤΙΩΜΕΝΗ_ΚΑΤΑΝΟΜΗ.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        type="primary",
                        use_container_width=True
                    )
                    
                    st.balloons()
                    
                except Exception as e:
                    st.error(f"❌ Σφάλμα Phase 2: {str(e)}")
                    with st.expander("Λεπτομέρειες"):
                        import traceback
                        st.code(traceback.format_exc())
    else:
        st.info("👆 Ανέβασε και τα δύο αρχεία")
    
    st.markdown("---")
    st.success("✅ v3.9 FINAL | Using .startswith() from working code")


if __name__ == '__main__':
    main()